import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows



df = pd.read_excel("Result2.xlsx")  # read the data from excel

df['video_link'] = df['video_id'].apply(lambda x: 'https://www.youtube.com/watch?v=' + str(x))  # prepend 'VID_' to every id in the 'ID' column
df = df.drop_duplicates(subset='video_id')

# Define the highlight style
# Initialize a Workbook
wb = Workbook()
ws = wb.active

# Convert dataframe to rows
rows = dataframe_to_rows(df, index=False, header=True)

highlight_fill = PatternFill(fill_type="solid", fgColor="FFFF00")

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)

        # Check if the column name 'tag' contains "News" or "TV"
        if c_idx == df.columns.get_loc('tags') + 1 and ('News' in str(value) or 'TV' in str(value)):
            # Highlight the whole row
            for col in range(1, len(row) + 1):
                ws.cell(row=r_idx, column=col).fill = highlight_fill

# Save the workbook
wb.save("highlighted_file.xlsx")


#df.to_excel("finalResult2.xlsx", index=False)  # write the data back to excel
